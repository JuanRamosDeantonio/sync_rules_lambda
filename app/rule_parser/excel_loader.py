from typing import List, Optional, Dict, Any
import time
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from app.utils.logger import get_logger

try:
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    InvalidFileException = Exception

logger = get_logger(__name__)

class ExcelProcessingError(Exception):
    def __init__(self, message: str, file_path: Optional[str] = None, sheet_name: Optional[str] = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        super().__init__(message)

class ExcelFileNotFoundError(ExcelProcessingError):
    pass

class InvalidExcelFileError(ExcelProcessingError):
    pass

class HeaderNotFoundError(ExcelProcessingError):
    pass

class MissingRequiredFieldsError(ExcelProcessingError):
    def __init__(self, message: str, missing_fields: List[str], **kwargs):
        self.missing_fields = missing_fields
        super().__init__(message, **kwargs)

class RuleValidationError(ExcelProcessingError):
    def __init__(self, message: str, row_data: Dict[str, Any], row_number: int, **kwargs):
        self.row_data = row_data
        self.row_number = row_number
        super().__init__(message, **kwargs)

# Mapeo de campos del Excel a campos del JSON
EXCEL_FIELD_MAP = {
    "Id": "id",
    "Documentacion": "documentation",
    "Descripcion": "description",
    "Artefacto": "references",
    "Tipo": "type",
    "Criticidad": "criticality",
    "Tags": "explanation"
}

# Campos requeridos que deben estar presentes en el Excel
REQUIRED_EXCEL_FIELDS = list(EXCEL_FIELD_MAP.keys())

def load_rules_from_excel(excel_path: str, type_filter: Optional[str] = None) -> List[dict]:
    """
    Carga reglas desde un archivo Excel.
    
    Args:
        excel_path: Ruta al archivo Excel
        type_filter: Filtro por tipo de regla (None = todas, "SEMANTICA", "CONTENIDO")
    
    Returns:
        Lista de diccionarios con las reglas procesadas
    """
    start_time = time.time()
    file_path = Path(excel_path)

    logger.info("🚀 Iniciando carga de reglas", extra={
        "file_path": str(file_path),
        "file_name": file_path.name,
        "type_filter": type_filter or "TODOS LOS TIPOS (SIN FILTRO)"
    })

    if not file_path.exists():
        error_msg = f"Archivo no encontrado: {excel_path}"
        logger.error(f"❌ {error_msg}")
        raise ExcelFileNotFoundError(error_msg, file_path=excel_path)

    if file_path.suffix.lower() not in ['.xlsx', '.xlsm']:
        error_msg = f"Formato de archivo no soportado: {file_path.suffix}. Solo se admiten .xlsx y .xlsm"
        logger.error(f"❌ {error_msg}")
        raise InvalidExcelFileError(error_msg, file_path=excel_path)

    rules = []
    wb = None

    try:
        wb = load_workbook(excel_path, data_only=True)
        logger.info(f"📂 Archivo Excel cargado exitosamente. Hojas encontradas: {len(wb.sheetnames)}")
    except InvalidFileException as e:
        raise InvalidExcelFileError(f"Archivo Excel inválido: {e}", file_path=excel_path) from e
    except Exception as e:
        raise InvalidExcelFileError(f"Error inesperado al abrir archivo Excel: {e}", file_path=excel_path) from e

    try:
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                logger.info(f"📄 Procesando hoja: {sheet_name}")
                sheet_rules = _process_worksheet(ws, sheet_name, type_filter)
                rules.extend(sheet_rules)
            except (HeaderNotFoundError, MissingRequiredFieldsError) as e:
                logger.warning(f"⚠️ Hoja {sheet_name} omitida: {e}")
                continue
            except Exception as e:
                logger.error(f"❌ Error inesperado procesando hoja {sheet_name}: {e}", exc_info=True)
                continue
    finally:
        if wb:
            wb.close()

    logger.info(f"✅ Proceso completado en {time.time() - start_time:.2f}s", extra={"total_rules": len(rules)})
    return rules

def _process_worksheet(ws: Worksheet, sheet_name: str, type_filter: Optional[str]) -> List[dict]:
    """Procesa una hoja de trabajo específica del Excel."""
    rules = []
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise HeaderNotFoundError("Hoja vacía", sheet_name=sheet_name)

    header_row_idx = _find_header_row_index(rows)
    if header_row_idx is None:
        raise HeaderNotFoundError("No se encontró encabezado con campos requeridos", sheet_name=sheet_name)

    # Determinar columnas válidas
    max_columns = _determine_max_columns(rows, header_row_idx)
    
    # Mapear headers
    mapped_headers = _get_mapped_headers(rows[header_row_idx], max_columns)
    
    # Crear template base
    base_template = _create_base_template(mapped_headers)

    logger.info(f"📊 Hoja '{sheet_name}': {max_columns} columnas, template con {len(base_template)} campos")

    # Procesar filas
    processed_rows = 0
    rules_with_references = 0
    
    for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        if _is_empty_row(row):
            continue
            
        try:
            row_data = _create_row_dict(row, mapped_headers, base_template)
            processed_rows += 1
            
            # Verificar references
            if row_data.get("references") is not None:
                rules_with_references += 1
                logger.info(f"🎯 ID {row_data.get('id')}: references='{row_data.get('references')}'")
            
            # Aplicar filtro si existe
            if type_filter and not _matches_type_filter(row_data, type_filter):
                continue
                
            rules.append(row_data)
            
        except Exception as e:
            logger.error(f"Error procesando fila {row_idx}: {e}")
    
    logger.info(f"✅ Hoja '{sheet_name}': {processed_rows} filas procesadas, {rules_with_references} con references, {len(rules)} incluidas")
    return rules

def _determine_max_columns(rows: List[tuple], header_row_idx: int) -> int:
    """Determina el número máximo de columnas con headers válidos."""
    header_row = rows[header_row_idx] if header_row_idx < len(rows) else ()
    max_cols = 0
    
    if header_row:
        for i, header in enumerate(header_row):
            if header and str(header).strip():
                max_cols = i + 1
    
    return max_cols

def _get_mapped_headers(header_row: tuple, max_columns: int) -> List[str]:
    """Mapea los headers del Excel a nombres del JSON."""
    if not header_row:
        header_row = ()
    
    mapped_headers = []
    for i in range(max_columns):
        if i < len(header_row):
            header = header_row[i]
            if header and str(header).strip():
                header_str = str(header).strip()
                mapped_header = EXCEL_FIELD_MAP.get(header_str, header_str)
                mapped_headers.append(mapped_header)
    
    return mapped_headers

def _create_base_template(mapped_headers: List[str]) -> Dict[str, Any]:
    """Crea template base con todos los campos."""
    template = {}
    
    # Incluir headers mapeados
    for header in mapped_headers:
        template[header] = None
    
    # Asegurar campos obligatorios
    for original_field, mapped_field in EXCEL_FIELD_MAP.items():
        if mapped_field not in template:
            template[mapped_field] = None
    
    return template

def _create_row_dict(row: tuple, mapped_headers: List[str], base_template: Dict[str, Any]) -> Dict[str, Any]:
    """Convierte una fila del Excel a diccionario."""
    row_dict = base_template.copy()
    
    if not row:
        return row_dict
    
    padded_row = list(row)
    while len(padded_row) < len(mapped_headers):
        padded_row.append(None)
    
    for i, header in enumerate(mapped_headers):
        if i < len(padded_row):
            value = padded_row[i]
            if isinstance(value, str):
                value = value.strip()
                row_dict[header] = value if value else None
            else:
                row_dict[header] = value
    
    return row_dict

def _matches_type_filter(row_dict: Dict[str, Any], type_filter: str) -> bool:
    """Verifica si una fila coincide con el filtro de tipo."""
    if not type_filter:
        return True
    
    row_type = row_dict.get("type")
    if row_type is None:
        return True  # Incluir filas con tipo None
    
    return str(row_type).strip().upper() == type_filter.upper()

def _is_empty_row(row: tuple) -> bool:
    """Verifica si una fila está completamente vacía."""
    return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)

def _find_header_row_index(rows: List[tuple]) -> Optional[int]:
    """Encuentra el índice de la fila que contiene los headers."""
    required_set = set(REQUIRED_EXCEL_FIELDS)
    for idx, row in enumerate(rows):
        if row:
            cell_values = set(str(cell).strip() for cell in row if cell)
            if required_set.issubset(cell_values):
                return idx
    return None